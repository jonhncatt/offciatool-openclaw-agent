from __future__ import annotations

import base64
import io
import re
import zipfile
from html import unescape
from pathlib import Path

_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_MSG_MARKERS_ASCII = (b"__substg1.0_", b"IPM.")
_MSG_MARKERS_UTF16 = tuple(marker.decode("ascii").encode("utf-16-le") for marker in _MSG_MARKERS_ASCII)
_XLSX_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _truncate(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    keep = max_chars - 64
    return f"{text[:keep]}\n\n[内容已截断，原始长度 {len(text)} 字符]"


def _read_plain_text(path: Path, max_chars: int) -> str:
    text = path.read_text(encoding="utf-8", errors="ignore")
    return _truncate(text, max_chars)


def _extract_pdf(path: Path, max_chars: int) -> str:
    from pypdf import PdfReader  # lazy import

    reader = PdfReader(str(path))
    chunks: list[str] = []
    for idx, page in enumerate(reader.pages, start=1):
        chunks.append(f"\n--- Page {idx} ---\n")
        chunks.append(page.extract_text() or "")
        if sum(len(c) for c in chunks) > max_chars:
            break
    return _truncate("".join(chunks), max_chars)


def _extract_docx(path: Path, max_chars: int) -> str:
    from docx import Document  # lazy import

    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return _truncate(text, max_chars)


def _xlsx_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        try:
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
        except Exception:
            pass
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return str(value.isoformat())
        except Exception:
            pass
    return str(value).strip()


def looks_like_xlsx_file(path: Path) -> bool:
    try:
        if not zipfile.is_zipfile(path):
            return False
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
        return "xl/workbook.xml" in names
    except Exception:
        return False


def _extract_xlsx(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook  # lazy import
    except Exception as exc:
        raise RuntimeError(
            "解析 .xlsx 需要依赖 openpyxl。请执行 `pip install -r requirements.txt` 后重试。"
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        lines: list[str] = ["[Excel 工作簿解析]"]
        total_chars = len(lines[0])
        truncated = False
        for sheet in wb.worksheets:
            title = (sheet.title or "").strip() or "Sheet"
            sheet_header = f"\n--- Sheet: {title} ---"
            lines.append(sheet_header)
            total_chars += len(sheet_header)
            if total_chars >= max_chars:
                truncated = True
                break

            sheet_rows = 0
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [_xlsx_cell_to_text(cell) for cell in row]
                while cells and not cells[-1]:
                    cells.pop()
                if not cells or not any(cells):
                    continue

                row_line = f"{row_idx}: " + " | ".join(cells)
                lines.append(row_line)
                total_chars += len(row_line)
                sheet_rows += 1
                if total_chars >= max_chars:
                    truncated = True
                    break

            if sheet_rows == 0:
                empty_line = "[空表或无可读内容]"
                lines.append(empty_line)
                total_chars += len(empty_line)
            if truncated:
                break

        if truncated:
            lines.append("\n[内容已截断，工作簿内容较大]")
        return _truncate("\n".join(lines), max_chars)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _html_to_text(html: str) -> str:
    raw = html or ""
    raw = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
    raw = re.sub(r"(?i)<br\\s*/?>", "\n", raw)
    raw = re.sub(r"(?i)</(p|div|li|tr|h1|h2|h3|h4|h5|h6|section|article)>", "\n", raw)
    raw = re.sub(r"(?s)<[^>]+>", " ", raw)
    raw = unescape(raw)
    lines: list[str] = []
    for line in raw.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip()
        if normalized:
            lines.append(normalized)
    return "\n".join(lines)


def _decode_bytes_best_effort(raw: bytes) -> str:
    if not raw:
        return ""
    for encoding in ("utf-8", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            out = raw.decode(encoding, errors="ignore")
        except Exception:
            continue
        if out.strip():
            return out
    return raw.decode("utf-8", errors="ignore")


def _looks_binaryish_text(text: str) -> bool:
    if not text:
        return False
    sample = text[:4096]
    if not sample:
        return False

    bad = 0
    for ch in sample:
        code = ord(ch)
        if code == 0:
            bad += 3
        elif code < 32 and ch not in "\n\r\t":
            bad += 1

    ratio = bad / max(1, len(sample))
    return ratio >= 0.02


def looks_like_outlook_msg_bytes(raw: bytes) -> bool:
    if not raw or not raw.startswith(_OLE2_MAGIC):
        return False
    head = raw[: max(4096, min(len(raw), 512 * 1024))]
    if any(marker in head for marker in _MSG_MARKERS_ASCII):
        return True
    if any(marker in head for marker in _MSG_MARKERS_UTF16):
        return True
    return False


def looks_like_outlook_msg_file(path: Path) -> bool:
    try:
        with path.open("rb") as fp:
            head = fp.read(512 * 1024)
    except Exception:
        return False
    return looks_like_outlook_msg_bytes(head)


def _extract_msg_body(msg: object) -> str:
    body = ""
    try:
        plain = getattr(msg, "body", None)
        if isinstance(plain, str):
            body = plain.strip()
        elif isinstance(plain, (bytes, bytearray)):
            body = _decode_bytes_best_effort(bytes(plain)).strip()
    except Exception:
        body = ""
    if body and not _looks_binaryish_text(body):
        return body

    try:
        html_body = getattr(msg, "htmlBody", None)
        if isinstance(html_body, (bytes, bytearray)):
            html_body = _decode_bytes_best_effort(bytes(html_body))
        if isinstance(html_body, str) and html_body.strip():
            html_text = _html_to_text(html_body).strip()
            if html_text and not _looks_binaryish_text(html_text):
                return html_text
    except Exception:
        pass

    try:
        rtf_body = getattr(msg, "rtfBody", None)
        deencap = getattr(msg, "deencapsulateBody", None)
        if rtf_body and callable(deencap):
            try:
                from extract_msg.enums import DeencapType  # lazy import

                plain_rtf = deencap(rtf_body, DeencapType.PLAIN)
            except Exception:
                plain_rtf = None
            if isinstance(plain_rtf, (bytes, bytearray)):
                plain_rtf = _decode_bytes_best_effort(bytes(plain_rtf))
            if isinstance(plain_rtf, str):
                plain_rtf = plain_rtf.strip()
                if plain_rtf and not _looks_binaryish_text(plain_rtf):
                    return plain_rtf
    except Exception:
        pass

    return ""


def _format_msg_attachment_line(att: object, idx: int) -> str:
    name = (
        (getattr(att, "longFilename", None) or "")
        or (getattr(att, "filename", None) or "")
        or (getattr(att, "name", None) or "")
        or f"attachment_{idx}"
    )
    extras: list[str] = []

    att_type = str(getattr(att, "type", "") or "").strip()
    if att_type:
        extras.append(att_type.split(".")[-1].lower())

    mime = (getattr(att, "mimetype", None) or "").strip()
    if mime:
        extras.append(mime)

    data = None
    try:
        data = getattr(att, "data", None)
    except Exception:
        data = None

    if isinstance(data, (bytes, bytearray)):
        extras.append(f"{len(data)} bytes")
    else:
        nested_subject = (getattr(data, "subject", None) or "").strip() if data is not None else ""
        if nested_subject:
            extras.append(f"嵌套邮件: {nested_subject}")

    if extras:
        return f"- {name} ({', '.join(extras)})"
    return f"- {name}"


def _extract_outlook_msg(path: Path, max_chars: int) -> str:
    try:
        import extract_msg  # lazy import
    except Exception as exc:
        raise RuntimeError(
            "解析 .msg 需要依赖 extract-msg。请执行 `pip install -r requirements.txt` 后重试。"
        ) from exc

    msg = extract_msg.openMsg(str(path), strict=False, delayAttachments=False)
    try:
        subject = (msg.subject or "").strip()
        sender = (msg.sender or "").strip()
        to = (msg.to or "").strip()
        cc = (msg.cc or "").strip()
        date = str(msg.date or "").strip()
        class_type = str(getattr(msg, "classType", "") or "").strip()
        body = _extract_msg_body(msg)

        attachment_lines: list[str] = []
        for idx, att in enumerate(getattr(msg, "attachments", []) or [], start=1):
            attachment_lines.append(_format_msg_attachment_line(att, idx))

        sections: list[str] = ["[Outlook MSG 邮件解析]"]
        if class_type:
            sections.append(f"消息类型: {class_type}")
        if subject:
            sections.append(f"主题: {subject}")
        if sender:
            sections.append(f"发件人: {sender}")
        if to:
            sections.append(f"收件人: {to}")
        if cc:
            sections.append(f"抄送: {cc}")
        if date:
            sections.append(f"时间: {date}")
        if attachment_lines:
            sections.append("附件列表:")
            sections.extend(attachment_lines)
        if body:
            sections.append("\n--- 正文 ---\n")
            sections.append(body)
        else:
            sections.append("\n--- 正文 ---\n")
            sections.append("[未提取到可读正文：该邮件可能仅包含附件、图片或受限富文本内容]")

        return _truncate("\n".join(sections).strip(), max_chars)
    finally:
        close = getattr(msg, "close", None)
        if callable(close):
            try:
                close()
            except Exception:
                pass


def extract_document_text(path: str, max_chars: int) -> str | None:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    plain_suffixes = {
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".log",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".html",
        ".css",
        ".yaml",
        ".yml",
        ".xml",
    }

    try:
        if suffix in plain_suffixes:
            return _read_plain_text(file_path, max_chars)
        if suffix == ".pdf":
            return _extract_pdf(file_path, max_chars)
        if suffix == ".docx":
            return _extract_docx(file_path, max_chars)
        if suffix in _XLSX_SUFFIXES:
            return _extract_xlsx(file_path, max_chars)
        if suffix == ".xls":
            return "[暂不支持 .xls（二进制 Excel）直接解析，请先另存为 .xlsx 后再读取]"
        if suffix in {".zip", ".bin"} and looks_like_xlsx_file(file_path):
            return _extract_xlsx(file_path, max_chars)
        if suffix == ".msg" or looks_like_outlook_msg_file(file_path):
            return _extract_outlook_msg(file_path, max_chars)
    except Exception as exc:
        return f"[文档解析失败: {exc}]"

    return None


def _heic_to_jpeg_bytes(path: Path) -> bytes:
    try:
        from PIL import Image
        from pillow_heif import register_heif_opener

        register_heif_opener()
        image = Image.open(path)
        rgb = image.convert("RGB")
        buffer = io.BytesIO()
        rgb.save(buffer, format="JPEG", quality=92)
        return buffer.getvalue()
    except Exception as exc:
        raise RuntimeError(
            "HEIC/HEIF conversion requires pillow-heif. Please install dependencies from requirements.txt."
        ) from exc


def image_to_data_url_with_meta(path: str, mime: str) -> tuple[str, str | None]:
    """
    Returns (data_url, warning). For HEIC, fallback to original HEIC payload
    when local conversion is unavailable, so capable gateways can still consume it.
    """
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    raw: bytes
    out_mime = mime
    warning: str | None = None

    is_heic = suffix in {".heic", ".heif"} or mime in {"image/heic", "image/heif"}
    if is_heic:
        try:
            raw = _heic_to_jpeg_bytes(file_path)
            out_mime = "image/jpeg"
        except Exception:
            raw = file_path.read_bytes()
            out_mime = mime if mime.startswith("image/") else "image/heic"
            warning = "HEIC 未本地转码，已原始上传；若网关不支持 HEIC，请先转 JPG/PNG。"
    else:
        raw = file_path.read_bytes()

    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{out_mime};base64,{encoded}", warning


def image_to_data_url(path: str, mime: str) -> str:
    data_url, _ = image_to_data_url_with_meta(path, mime)
    return data_url


def summarize_file_payload(path: str, max_bytes: int = 768, max_text_chars: int = 1200) -> str:
    file_path = Path(path)
    raw = file_path.read_bytes()
    head = raw[:max_bytes]

    if not head:
        return "[空文件]"

    text_bytes = b"\n\r\t\b\f" + bytes(range(32, 127))
    non_text = sum(1 for b in head if b not in text_bytes)
    is_binary = b"\x00" in head or (non_text / len(head)) > 0.30

    if not is_binary:
        text = head.decode("utf-8", errors="ignore")
        text = text[:max_text_chars]
        return f"[文本预览，文件大小 {len(raw)} bytes]\\n{text}"

    hex_preview = " ".join(f"{b:02x}" for b in head[:128])
    return (
        f"[二进制预览，文件大小 {len(raw)} bytes，前 {min(len(head),128)} bytes(hex)]\\n"
        f"{hex_preview}"
    )
